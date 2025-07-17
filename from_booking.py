import fbextract
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

def from_booking_1(file_name,sheet_name = "Календар"):
    wb = load_workbook('xls/'+file_name)
    sheet = wb[sheet_name]
    # Зчитуємо дати з першого рядка
    dates = []
    for col in range(2, sheet.max_column + 1):
        cell_val = sheet.cell(row=1, column=col).value
        try:
            if isinstance(cell_val, datetime):
                date = cell_val
            else:
                date = datetime.strptime(str(cell_val), "%d.%m.%Y")
            dates.append(date)
        # except Exception as e:
        #     print(str(e))
        except:
            dates.append(None)

    con = fbextract.get_connection()
    cur = con.cursor()
    # Зчитуємо бронювання
    for row in range(2, sheet.max_row + 1):
        mil_num = sheet.cell(row=row, column=1).value
        if not mil_num:
            continue
        for col, date in enumerate(dates, start=2):
            order = sheet.cell(row=row, column=col).value
            if date and order:
                # Вставка в БД
                cur.execute(" execute procedure booking_upd (?,?,?) ", (mil_num, date.date(), order) )

    con.commit()
    con.close()

def from_booking(file_name):
    # 1. Зчитування таблиці з Excel
    df = pd.read_excel('tmp/'+file_name,usecols="C,N:T")
    # 2. Перетворення в рядки
    melted = df.melt(id_vars=['військовий номер'], var_name='DATE_', value_name='BAT_ORDER')
    removed = melted[melted["BAT_ORDER"].isna()]
    melted = melted.dropna(subset=['BAT_ORDER'])

    # 3. Підключення до Firebird
    con = fbextract.get_connection()
    cur = con.cursor()
    # 4. Вставка в таблицю
    for _, row in melted.iterrows():
        # print(row['MIL_NUM'],pd.to_datetime(row['DATE_']).date(),row['BAT_ORDER'])
        cur.execute(" execute procedure booking_upd (?,?,?) ",
                    (row['військовий номер'], pd.to_datetime(row['DATE_']).date(), row['BAT_ORDER']))
    for _, row in removed.iterrows():
        cur.execute(" execute procedure booking_upd (?,?,?) ",
                    (row['військовий номер'], pd.to_datetime(row['DATE_']).date(), 'null'))
    con.commit()
    con.close()

# test area
# from_booking('UPD_Автомобілі_1ЄБ.xlsx')