import fbextract
from openpyxl import load_workbook
from datetime import datetime, timedelta

con = fbextract.get_connection()
cur = con.cursor()

wb = load_workbook("xls/UPD_Автомобілі_1ЄБ.xlsx")
sheet = wb["Календар"]

# Зчитуємо дати з першого рядка
dates = []
for col in range(2, sheet.max_column + 1):
    cell_val = sheet.cell(row=1, column=col).value
    try:
        # Очікуємо формат типу "Пн 15.07"
        # date = datetime.strptime(cell_val.strip(), "%d.%m.%Y")
        if isinstance(cell_val, datetime):
            date = cell_val
        else:
            date = datetime.strptime(str(cell_val), "%d.%m.%Y")
        dates.append(date)
    # except Exception as e:
    #     print(str(e))
    except:
        dates.append(None)


# Зчитуємо бронювання
for row in range(2, sheet.max_row + 1):
    mil_num = sheet.cell(row=row, column=1).value
    if not mil_num:
        continue

    for col, date in enumerate(dates, start=2):
        order = sheet.cell(row=row, column=col).value
        print(order)
        # if date and order:
            # Вставка в БД
            # cur.execute("""
            #     INSERT INTO or update booking (car_id, date_, bat_order)
            #     VALUES (?, ?, ?)
            #
            # """, (mil_num, date.date(), order))

con.commit()
con.close()